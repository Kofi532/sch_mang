from django.shortcuts import render
from .models import report,report30,report70, reportn
from .forms import ReportForm
from uploading.models import fees_update
from users.models import use, sch_reg,act
import pandas as pd
import xlsxwriter
from django.http import FileResponse
import io
from datetime import date
import string
import openpyxl
from itertools import islice
from operator import add
import numpy as np


# Create your views here.

#['stu_id', 'subjectA', 'subjectB', 'subjectC', 'subjectD', 'subjectE', 'subjectF', 'subjectG', 'subjectH','subjectI','subjectJ', 'subjectK', 'subjectL',]
def download_sub(request):
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    username = None
    usernamed = request.user.username
    df_act = pd.DataFrame(act.objects.all().values().filter(username=usernamed))
    term = list(df_act['active_term']) 
    term = term[0]
    dfr = pd.DataFrame(sch_reg.objects.all().values().filter(username=usernamed))
    if list(dfr['date']) == []:
        dfr = pd.DataFrame(use.objects.all().values().filter(username = usernamed))
    ffr = list(dfr['full_sch'])
    ffrc = list(dfr['contact_details'])
    schr = ffr[0]
    tel = ffrc[0]
    df = pd.DataFrame(use.objects.all().values().filter(username = usernamed))
    if list(df) == []:
        df = pd.DataFrame(sch_reg.objects.all().values().filter(username = usernamed))
    ff = list(df['school'])
    sch = ff[0] 
    today = str(date.today())
    # Create some cell formats with protection properties.
    unlocked = workbook.add_format({'locked': False})
    locked   = workbook.add_format({'locked': True})
    merge_format = workbook.add_format({
    'bold': 1,
    'border': 0,
    'align': 'center',
    'valign': 'vcenter'})

    merge_format1 = workbook.add_format({
    #'bold': 1,
    'border': 0,
    'align': 'center',
    'valign': 'vcenter'})

    f1= workbook.add_format()


    subjects = pd.DataFrame(reportn.objects.all().values().filter(school = sch).filter(stu_id = 'stu_id'))
    col = ['number','stu_id', 'subjectA', 'subjectB', 'subjectC', 'subjectD', 'subjectE', 'subjectF', 'subjectG', 'subjectH','subjectI','subjectJ', 'subjectK', 'subjectL']  
    subjects = subjects[col]
    #filter(stu_id = 'stu_id').values_list('number','subjectA', 'subjectB', 'subjectC', 'subjectD', 'subjectE', 'subjectF', 'subjectG', 'subjectH','subjectI','subjectJ', 'subjectK', 'subjectL')
    subjects['number'] = '###'
    subjects.columns = subjects.iloc[0]
    sub_list = subjects.columns
    sub_list = sub_list.insert(2, "Full Name")
    ree = ['Creche', 'Nursery1', 'Nursery2','K.G1', 'K.G2','Class1', 'Class2', 'Class3', 'Class4', 'Class5', 'Class6', 'J.H.S1', 'J.H.S2', 'J.H.S3']

    for t in ree:
        worksheet = workbook.add_worksheet(t+'-30%')
        worksheet2 = workbook.add_worksheet(t+'-70%')
        worksheet2.protect()
        worksheet.protect()

        row_num = 0 
        columns = sub_list
        for col_num in range(len(columns)):
            f1.set_bold(True)
            worksheet.write(row_num, col_num, columns[col_num], f1) 
            worksheet2.write(row_num, col_num, columns[col_num], f1)
          #  worksheet.write(row_num+1, col_num, 0, f1)
            worksheet.set_column(row_num, col_num, 13)
            worksheet2.set_column(row_num, col_num, 13)

        df = pd.DataFrame(fees_update.objects.all().values().filter(school = sch).filter(level = t))
        if list(df) == []:
            df
        else:
            stu_id = list(df['stu_id'])
            f_name = list(df['firstname'])
            for m in f_name:
                f_name = list(map(lambda x: x.replace(m, ' '+ m+ ' '), f_name))
            m_name = list(df['middlename'])
            m_name = list(map(lambda x: x.replace('None', ''), m_name))
            l_name = list(df['lastname'])
            for m in l_name:
                l_name = list(map(lambda x: x.replace(m, ' '+ m+ ' '), l_name))            
            d = [i+j+k for i,j,k in zip(l_name,m_name,f_name)]
            for row in range(len(d)):
                worksheet.write(row+1, 2, d[row])
                worksheet2.write(row+1, 2, d[row])               
                worksheet.write(row+1, 1, stu_id[row])
                worksheet2.write(row+1, 1, stu_id[row])
                for col_num in range(len(columns)-3):
                    worksheet.write(row+1, col_num+3, 0, unlocked)
                    worksheet2.write(row+1, col_num+3, 0, unlocked)
                    alph = list(string.ascii_uppercase[3:len(columns)])
                    for j in alph:
                        place = j+str(row+1)
                        worksheet.data_validation(place, {'validate': 'decimal',
                                        'criteria': '<',
                                        'value': 31,
                                        'input_message': 'Please ensure cell contains only figures and it should be less than or equal to 30%'
                                        })
                        worksheet2.data_validation(place, {'validate': 'decimal',
                                        'criteria': '<',
                                        'value': 71,
                                        'input_message': 'Please ensure cell contains only figures and it should be less than or equal to 70%'
                                        })
    workbook.close()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename='reports.xlsx')

def upload_report(request):
    if "GET" == request.method:
        return render(request, 'report.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        username = None
        usernamed = request.user.username 
        # getting a particular sheet by name out of many sheets
        ree = ['Creche', 'Nursery1', 'Nursery2', 'K.G1', 'K.G2','Class1', 'Class2', 'Class3', 'Class4', 'Class5', 'Class6', 'J.H.S1', 'J.H.S2', 'J.H.S3']
        #ree = ['K.G1']
        dfs = pd.DataFrame(use.objects.all().values().filter(username = usernamed))
        if list(dfs) == []:
            dfs = pd.DataFrame(sch_reg.objects.all().values().filter(username = usernamed))
        code = list(dfs['school'])
        code = code[0]
        df_act = pd.DataFrame(act.objects.all().values())
        df_act = df_act[df_act['school_code'] == code]
        term = list(df_act['active_term']) 
        term = term[0]

        for i in ree:
            worksheet = wb[i+'-30%']
            worksheet2 = wb[i+'-70%']
            data = worksheet.values
            data2 = worksheet2.values
            cols = next(data)[1:]
            cols2 = next(data2)[1:]
            data = list(data)
            data2 = list(data2)
            idx = [r[0] for r in data]
            idx2 = [r[0] for r in data2]
            data = (islice(r, 1, None) for r in data)
            data2 = (islice(r, 1, None) for r in data2)
            df = pd.DataFrame(data, index=idx, columns=cols)
            df2 = pd.DataFrame(data2, index=idx2, columns=cols2)
            df = df.drop(['Full Name'], axis=1)
            df2 = df2.drop(['Full Name'], axis=1)
            df.insert(0, "number", list(df['stu_id']), True)
            df2.insert(0, "number", list(df2['stu_id']), True)
            df['school'] = code
            df2['school'] = code
            df['level'] = i
            df2['level'] = i 
            c_list = list(df.columns)
            c_list2 = list(df2.columns)
            chief_ = ['number', 'stu_id', 'subjectA', 'subjectB', 'subjectC', 'subjectD', 'subjectE', 'subjectF', 'subjectG', 'subjectH','subjectI','subjectJ', 'subjectK', 'subjectL','school', 'level']
            chief = chief_[:len(c_list)]
            chief2 = chief_[:len(c_list2)]
            df.set_axis(chief, axis='columns', inplace=True)
            df2.set_axis(chief2, axis='columns', inplace=True)
            for index, row in df.iterrows():
                model = report30()
                model.number = row['number']
                model.stu_id = row['stu_id']
                model.subjectA = row['subjectA']
                model.subjectB = row['subjectB']
                model.subjectC = row['subjectC']
                model.subjectD = row['subjectD']
                model.subjectE = row['subjectE']
                model.subjectF = row['subjectF']
                model.subjectG = row['subjectG']
                model.subjectH = row['subjectH']
                model.subjectI = row['subjectI']
                model.subjectJ = row['subjectJ']
                model.subjectK = row['subjectK']
                model.subjectL = row['subjectL']
                model.school = row['school']
                model.level = row['level']
                model.save()
            for index, row in df2.iterrows():
                model = report70()
                model.number = row['number']
                model.stu_id = row['stu_id']
                model.subjectA = row['subjectA']
                model.subjectB = row['subjectB']
                model.subjectC = row['subjectC']
                model.subjectD = row['subjectD']
                model.subjectE = row['subjectE']
                model.subjectF = row['subjectF']
                model.subjectG = row['subjectG']
                model.subjectH = row['subjectH']
                model.subjectI = row['subjectI']
                model.subjectJ = row['subjectJ']
                model.subjectK = row['subjectK']
                model.subjectL = row['subjectL']
                model.school = row['school']
                model.level = row['level']
                model.save()    
            df['subjectA'] = list( map(add, list(df['subjectA']), list(df2['subjectA'])) )
            df['subjectB'] = list( map(add, list(df['subjectB']), list(df2['subjectB'])) )
            df['subjectC'] = list( map(add, list(df['subjectC']), list(df2['subjectC'])) )
            df['subjectD'] = list( map(add, list(df['subjectD']), list(df2['subjectD'])) )
            df['subjectE'] = list( map(add, list(df['subjectE']), list(df2['subjectE'])) )
            df['subjectF'] = list( map(add, list(df['subjectF']), list(df2['subjectF'])) )
            df['subjectG'] = list( map(add, list(df['subjectG']), list(df2['subjectG'])) )
            df['subjectH'] = list( map(add, list(df['subjectH']), list(df2['subjectH'])) )
            df['subjectI'] = list( map(add, list(df['subjectI']), list(df2['subjectI'])) )
            df['subjectJ'] = list( map(add, list(df['subjectJ']), list(df2['subjectJ'])) )
            df['subjectK'] = list( map(add, list(df['subjectK']), list(df2['subjectK'])) )
            df['subjectL'] = list( map(add, list(df['subjectL']), list(df2['subjectL'])) )

            for index, row in df.iterrows():
                model = report()
                model.number = row['number']
                model.stu_id = row['stu_id']
                model.subjectA = row['subjectA']
                model.subjectB = row['subjectB']
                model.subjectC = row['subjectC']
                model.subjectD = row['subjectD']
                model.subjectE = row['subjectE']
                model.subjectF = row['subjectF']
                model.subjectG = row['subjectG']
                model.subjectH = row['subjectH']
                model.subjectI = row['subjectI']
                model.subjectJ = row['subjectJ']
                model.subjectK = row['subjectK']
                model.subjectL = row['subjectL']
                model.school = row['school']
                model.level = row['level']
                model.save()
        return render(request, 'upload.html', {})




def report_reg(request):
    form = ReportForm(request.POST or None)
    username = None
    usernamed = request.user.username 
    dfs = pd.DataFrame(use.objects.all().values().filter(username = usernamed))
    if list(dfs) == []:
        dfs = pd.DataFrame(sch_reg.objects.all().values().filter(username = usernamed))
    ff = list(dfs['school'])
    sch = ff[0]
    level = 'Creche'
    if request.method == 'POST'and form.is_valid():
        subjectA = request.POST.get('subjectA')
        subjectB = request.POST.get('subjectB')
        subjectC = request.POST.get('subjectC')
        subjectD = request.POST.get('subjectD')
        subjectE = request.POST.get('subjectE')
        subjectF = request.POST.get('subjectF')
        subjectG = request.POST.get('subjectG')
        subjectH = request.POST.get('subjectH')
        subjectI = request.POST.get('subjectI')
        subjectJ = request.POST.get('subjectJ')
        subjectK = request.POST.get('subjectK')
        subjectL = request.POST.get('subjectL')
        mod =reportn(level = level,school = sch, stu_id = 'stu_id', subjectA=subjectA, subjectB=subjectB, subjectC=subjectC,subjectD=subjectD, subjectE=subjectE, subjectF=subjectF, subjectG=subjectG, subjectH=subjectH, subjectI=subjectI, subjectJ=subjectJ, subjectK=subjectK, subjectL=subjectL )
        mod.save()
 #       thanks = 'Subjects added as follows'
            
#rows = fees_update.objects.all().filter(school = sch).filter(level = t).values_list('stu_id', 'firstname' , 'middlename', 'lastname', 'level', 'fee','balance', 'amountpaid_term1', 'amountpaid_term2','amountpaid_term3' )
    df = pd.DataFrame(reportn.objects.all().filter(school = sch).filter(stu_id = 'stu_id'))
    if list(df) == []:
        return render(request, 'addsubject.html',{'form': form})
    else: 
        return render (request, 'correct.html', {})


def report_cards(request):
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    username = None
    usernamed = request.user.username
    df_act = pd.DataFrame(act.objects.all().values().filter(username=usernamed))
    term = list(df_act['active_term']) 
    term = term[0]
    dfr = pd.DataFrame(sch_reg.objects.all().values().filter(username=usernamed))
    if list(dfr['date']) == []:
        dfr = pd.DataFrame(use.objects.all().values().filter(username = usernamed))
    ffr = list(dfr['full_sch'])
    ffrc = list(dfr['contact_details'])
    schr = ffr[0]
    tel = ffrc[0]
    df = pd.DataFrame(use.objects.all().values().filter(username = usernamed))
    if list(df) == []:
        df = pd.DataFrame(sch_reg.objects.all().values().filter(username = usernamed))
    ff = list(df['school'])
    sch = ff[0] 
    today = str(date.today())
    # Create some cell formats with protection properties.
    unlocked = workbook.add_format({'locked': False})
    locked   = workbook.add_format({'locked': True})
    merge_format = workbook.add_format({
    'bold': 1,
    'border': 0,
    'align': 'center',
    'valign': 'vcenter'})

    merge_format1 = workbook.add_format({
    #'bold': 1,
    'border': 0,
    'align': 'center',
    'valign': 'vcenter'})

    f1= workbook.add_format()
    ree = ['Creche']

    for i in ree:
        df = pd.DataFrame(report.objects.all().values().filter(school=sch).filter(level = i))
        students = list(df['stu_id'])
        for q in students:
            worksheet = workbook.add_worksheet(q)
            worksheet.protect()
            worksheet.merge_range('A1:E1', schr, merge_format)
            worksheet.merge_range('A2:E2', 'Tel: '+tel , merge_format)
            columns1 = ['Stu_id' , 'Firstname' , 'Middlename', 'Lastname', 'Level']
            alp = list(string.ascii_uppercase[2:len(columns1)])
            df = pd.DataFrame(fees_update.objects.all().values().filter(school=sch).filter(level = i).filter(stu_id = q))
            if list(df) == []:
                df
            else:
                columns2 = ['stu_id' , 'firstname' , 'middlename', 'lastname', 'level']
                df_in = df[columns2]
                df_list = list(df_in.iloc[0])
                row_num = 4
                for col_num in range(len(df_list)): ##add real name
                    worksheet.write(row_num, col_num, df_list[col_num])
            row_num = 3
            for col_num in range(len(columns1)):
                worksheet.write(row_num, col_num, columns1[col_num], merge_format)  #adding name thems columns
                #worksheet.write(row_num+1, col_num, columns1[col_num], merge_format)
                worksheet.set_column(row_num, col_num, 20)
            df = pd.DataFrame(reportn.objects.all().values().filter(school=sch).filter(level = i))
            cols = ['subjectA', 'subjectB', 'subjectC', 'subjectD', 'subjectE', 'subjectF', 'subjectG', 'subjectH','subjectI','subjectJ', 'subjectK', 'subjectL']
            df = df[cols]
            lists = list(df.iloc[0])
            lists = list(dict.fromkeys(lists))
            lists.remove("0")
            lists.insert(0, "id")
            length = len(lists)
            pee = df.columns
            for row_num in range(len(lists)): ##Subject lists
                col_num = 0
                worksheet.write(row_num+6, col_num, lists[row_num]) 
                worksheet.set_column(row_num+6, col_num, 20)
            worksheet.write(6, 0, 'Subjects', merge_format) ##Subjects
            worksheet.write(6, 1, '30%', merge_format)
            worksheet.write(6, 2, '70%', merge_format)
            worksheet.write(6, 3, 'Total', merge_format)
            worksheet.write(6, 4, 'Class Position', merge_format)
            df30 = pd.DataFrame(report30.objects.all().values().filter(school=sch).filter(level = i).filter(stu_id=q)) 
            df30 = df30[cols]
            lists30 = list(df30.iloc[0])
         #   lists30= list(dict.fromkeys(lists30))
         #   lists.remove("0")
         #   lists.insert(0, "id")
            lists30 = lists30[0:length-1]
            for row_num in range(len(lists30)): ##30% marks
                col_num = 1
                worksheet.write(row_num+7, col_num, lists30[row_num]) 
            df70 = pd.DataFrame(report70.objects.all().values().filter(school=sch).filter(level = i).filter(stu_id=q)) 
            df70 = df70[cols]
            lists70 = list(df70.iloc[0])
         #   lists30= list(dict.fromkeys(lists30))
         #   lists.remove("0")
         #   lists.insert(0, "id")
            lists70 = lists70[0:length-1]
            for row_num in range(len(lists70)): ##70% marks
                col_num = 2
                worksheet.write(row_num+7, col_num, lists70[row_num]) 
            df100 = pd.DataFrame(report.objects.all().values().filter(school=sch).filter(level = i).filter(stu_id=q)) 
            df100 = df100[cols]
            lists100 = list(df100.iloc[0])
         #   lists30= list(dict.fromkeys(lists30))
         #   lists.remove("0")
         #   lists.insert(0, "id")
            lists100 = lists100[0:length-1]
            for row_num in range(len(lists100)): ##100% marks
                col_num = 3
                worksheet.write(row_num+7, col_num, lists100[row_num]) 
            dft = pd.DataFrame(report.objects.all().values().filter(school=sch).filter(level = i)) 
            studs = list(dft['stu_id'])
            dfr = pd.DataFrame(report.objects.all().values().filter(school=sch).filter(level = i).filter(stu_id = q))
            person = list(dfr['stu_id'])[0] 
            
            dft = dft[cols]
            for p in cols:
                lis = list(dft[p])
                pos = [sorted(lis).index(x) for x in lis]
                ag = len(pos)
                pos = [ag-x for x in pos]          
                #pos = [x+1 for x in pos]
                dft[p] = pos
                dft['stu_id'] = studs
                dftt = dft[dft['stu_id']==person]
                listn = list(dftt.iloc[0])
                listn = listn[0:length-1]
            for row_num in range(len(listn)): ##100% marks
                col_num = 4
                worksheet.write(row_num+7, col_num, listn[row_num])           
            
            

            
               # df100 = df100.sort_values([p], ascending=[True])
     #       for z in cols:
                




            #df100.sort_values('')
            
             

    workbook.close()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename='reportcards.xlsx')